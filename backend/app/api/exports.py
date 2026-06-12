"""Export-Endpoints: PDF und Excel."""
from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from sqlalchemy.orm import Session

from app.models.database import get_db
from app.models.models import Insurance, Product

router = APIRouter()

# Zeichen, die Excel als Formel-Beginn interpretiert (Formula Injection)
_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _append_text_row(ws, values: list) -> None:
    """Fügt eine Zeile an und erzwingt Text-Zellen für formelartige Strings.

    openpyxl interpretiert Strings mit führendem '=' als Formel — Werte aus
    Nutzereingaben (Name, Vertragsnummer, Notizen) dürfen nie als Formel landen.
    """
    ws.append(values)
    for cell in ws[ws.max_row]:
        if isinstance(cell.value, str) and cell.value.startswith(_FORMULA_PREFIXES):
            cell.data_type = "s"


@router.get("/insurances.xlsx")
def export_insurances_xlsx(db: Session = Depends(get_db)) -> StreamingResponse:
    rows = db.query(Insurance).all()
    wb = Workbook()
    ws = wb.active
    ws.title = "Versicherungen"
    ws.append(
        [
            "Name",
            "Kategorie",
            "Versicherer",
            "Vertragsnummer",
            "Start",
            "Ende",
            "Prämie EUR",
            "Intervall",
            "Notizen",
        ]
    )
    for r in rows:
        _append_text_row(
            ws,
            [
                r.name,
                r.kategorie.value,
                r.versicherer,
                r.vertragsnummer,
                r.start_date.isoformat() if r.start_date else "",
                r.end_date.isoformat() if r.end_date else "",
                r.praemie_eur or 0,
                r.zahlungsintervall.value,
                r.notes or "",
            ],
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="versicherungen.xlsx"'},
    )


@router.get("/insurances.pdf")
def export_insurances_pdf(db: Session = Depends(get_db)) -> StreamingResponse:
    rows = db.query(Insurance).all()
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title="Versicherungsübersicht")
    styles = getSampleStyleSheet()
    elements = [Paragraph("Versicherungsübersicht", styles["Title"]), Spacer(1, 12)]

    data = [["Name", "Kategorie", "Versicherer", "Ende", "Prämie €", "Intervall"]]
    for r in rows:
        data.append(
            [
                r.name,
                r.kategorie.value,
                r.versicherer,
                r.end_date.isoformat() if r.end_date else "-",
                f"{r.praemie_eur:.2f}" if r.praemie_eur is not None else "-",
                r.zahlungsintervall.value,
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1976d2")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5f5")]),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/pdf",
        headers={"Content-Disposition": 'attachment; filename="versicherungen.pdf"'},
    )


@router.get("/products.xlsx")
def export_products_xlsx(db: Session = Depends(get_db)) -> StreamingResponse:
    rows = db.query(Product).all()

    # Build insurance name lookup to avoid raw IDs in the export
    ins_map: dict[int, str] = {
        i.id: f"{i.name} ({i.versicherer})"
        for i in db.query(Insurance).all()
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Produkte"
    ws.append(["Name", "Kategorie", "Kaufdatum", "Garantieende", "Verknüpfte Versicherung", "Notizen"])
    for r in rows:
        _append_text_row(
            ws,
            [
                r.name,
                r.kategorie,
                r.purchase_date.isoformat() if r.purchase_date else "",
                r.warranty_end.isoformat() if r.warranty_end else "",
                ins_map.get(r.linked_insurance_id, "") if r.linked_insurance_id else "",
                r.notes or "",
            ],
        )
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="produkte.xlsx"'},
    )
