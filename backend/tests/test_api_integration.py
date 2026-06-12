"""Integrations-Test aller API-Endpoints."""
from __future__ import annotations

from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook

from app.main import app

client = TestClient(app)


@pytest.fixture(scope="module", autouse=True)
def _setup_db() -> None:
    from app.models.database import init_db

    init_db()


def test_health() -> None:
    r = client.get("/api/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok"}


def test_list_insurances_empty() -> None:
    r = client.get("/api/insurances")
    assert r.status_code == 200
    assert isinstance(r.json(), list)


_INSURANCE_PAYLOAD = {
    "name": "Test KFZ",
    "kategorie": "KFZ",
    "versicherer": "Allianz",
    "vertragsnummer": "KFZ-123",
    "start_date": "2025-01-01",
    "end_date": "2026-12-31",
    "praemie_eur": 600,
    "zahlungsintervall": "jährlich",
}


def test_create_and_get_insurance() -> None:
    r = client.post("/api/insurances", json=_INSURANCE_PAYLOAD)
    assert r.status_code == 201
    data = r.json()
    assert data["id"]
    ins_id = data["id"]

    r = client.get(f"/api/insurances/{ins_id}")
    assert r.status_code == 200
    assert r.json()["name"] == "Test KFZ"


def test_insurance_with_kuendigung_fields() -> None:
    """Kündigung als zwei wiederkehrende Daten (Tag+Monat ohne Jahr) speichern und lesen."""
    payload = {
        **_INSURANCE_PAYLOAD,
        "kuendigung_bis_tag": 30,
        "kuendigung_bis_monat": 9,
        "kuendigung_zum_tag": 31,
        "kuendigung_zum_monat": 12,
    }
    r = client.post("/api/insurances", json=payload)
    assert r.status_code == 201
    data = r.json()
    assert data["kuendigung_bis_tag"] == 30
    assert data["kuendigung_bis_monat"] == 9
    assert data["kuendigung_zum_tag"] == 31
    assert data["kuendigung_zum_monat"] == 12
    client.delete(f"/api/insurances/{data['id']}")


def test_insurance_kuendigung_validation() -> None:
    # Tag ohne Monat → abgelehnt
    r = client.post(
        "/api/insurances",
        json={**_INSURANCE_PAYLOAD, "kuendigung_bis_tag": 15},
    )
    assert r.status_code == 422

    # 31. Februar existiert nicht → abgelehnt
    r = client.post(
        "/api/insurances",
        json={**_INSURANCE_PAYLOAD, "kuendigung_zum_tag": 31, "kuendigung_zum_monat": 2},
    )
    assert r.status_code == 422


def test_financial_summary() -> None:
    r = client.get("/api/insurances/summary/financial")
    assert r.status_code == 200
    body = r.json()
    assert "total_year_eur" in body
    assert "total_month_eur" in body
    assert "by_category" in body


def test_create_product() -> None:
    pdata = {
        "name": "MacBook Pro",
        "kategorie": "Elektronik",
        "purchase_date": "2025-06-01",
        "warranty_end": "2027-06-01",
    }
    r = client.post("/api/products", json=pdata)
    assert r.status_code == 201
    assert r.json()["id"]


def test_warranty_status() -> None:
    r = client.get("/api/products/summary/warranty-status")
    assert r.status_code == 200
    body = r.json()
    for key in ("green", "yellow", "red", "expired", "no_warranty"):
        assert key in body


def test_export_pdf() -> None:
    r = client.get("/api/exports/insurances.pdf")
    assert r.status_code == 200
    assert r.headers["content-type"] == "application/pdf"
    assert len(r.content) > 0


def test_export_xlsx() -> None:
    r = client.get("/api/exports/insurances.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert len(r.content) > 0


def test_export_products_xlsx() -> None:
    r = client.get("/api/exports/products.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]


def test_update_insurance() -> None:
    r = client.post("/api/insurances", json=_INSURANCE_PAYLOAD)
    assert r.status_code == 201
    ins_id = r.json()["id"]

    updated = {**_INSURANCE_PAYLOAD, "name": "Test KFZ Updated"}
    r = client.put(f"/api/insurances/{ins_id}", json=updated)
    assert r.status_code == 200
    assert r.json()["name"] == "Test KFZ Updated"


def test_delete_insurance() -> None:
    r = client.post("/api/insurances", json=_INSURANCE_PAYLOAD)
    assert r.status_code == 201
    ins_id = r.json()["id"]

    r = client.delete(f"/api/insurances/{ins_id}")
    assert r.status_code == 204

    r = client.get(f"/api/insurances/{ins_id}")
    assert r.status_code == 404


def test_invoice_lifecycle() -> None:
    # Produkt anlegen
    pdata = {
        "name": "Samsung TV",
        "kategorie": "Elektronik",
        "purchase_date": "2024-01-01",
        "warranty_end": "2027-01-01",  # 3-jährige Garantie
    }
    r = client.post("/api/products", json=pdata)
    assert r.status_code == 201
    product_id = r.json()["id"]

    # Rechnung hochladen (minimales PDF-ähnliches PNG)
    png_bytes = (
        b"\x89PNG\r\n\x1a\n"
        + b"\x00" * 100  # dummy content (16+ bytes, magic bytes korrekt)
    )
    r = client.post(
        "/api/invoices",
        data={"product_id": product_id, "purchase_date": "2024-01-01", "amount_eur": "599.99"},
        files={"file": ("rechnung.png", png_bytes, "image/png")},
    )
    assert r.status_code == 201
    inv = r.json()
    assert inv["product_id"] == product_id
    assert inv["amount_eur"] == 599.99
    # retain_until muss = warranty_end sein, da warranty_end > purchase_date + 730 Tage
    assert inv["retain_until"] == "2027-01-01"

    inv_id = inv["id"]

    # Rechnung abrufen
    r = client.get(f"/api/invoices/{inv_id}")
    assert r.status_code == 200

    # Liste abrufen
    r = client.get("/api/invoices")
    assert r.status_code == 200
    assert any(i["id"] == inv_id for i in r.json())

    # Liste nach Produkt filtern
    r = client.get(f"/api/invoices?product_id={product_id}")
    assert r.status_code == 200
    assert all(i["product_id"] == product_id for i in r.json())

    # Löschen soll abgelehnt werden (Frist noch aktiv)
    r = client.delete(f"/api/invoices/{inv_id}")
    assert r.status_code == 409
    assert "Aufbewahrungsfrist" in r.json()["detail"]


def test_invoice_min_retention_without_warranty() -> None:
    """Ohne warranty_end muss die 2-Jahres-Mindestfrist gelten."""
    pdata = {"name": "Kaffeemaschine", "kategorie": "Haushaltsgeräte", "purchase_date": "2024-03-01"}
    r = client.post("/api/products", json=pdata)
    assert r.status_code == 201
    product_id = r.json()["id"]

    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    r = client.post(
        "/api/invoices",
        data={"product_id": product_id, "purchase_date": "2024-03-01"},
        files={"file": ("rechnung.png", png_bytes, "image/png")},
    )
    assert r.status_code == 201
    inv = r.json()
    # retain_until muss >= purchase_date + 730 Tage sein
    from datetime import date, timedelta
    expected = date(2024, 3, 1) + timedelta(days=730)
    assert inv["retain_until"] == expected.isoformat()


def test_invoice_product_not_found() -> None:
    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    r = client.post(
        "/api/invoices",
        data={"product_id": 999999},
        files={"file": ("rechnung.png", png_bytes, "image/png")},
    )
    assert r.status_code == 404


def test_delete_product_cascades_invoices() -> None:
    """Produkt löschen entfernt alle zugehörigen Rechnungen — auch in laufender Frist.

    Die Aufbewahrungsfrist blockiert nur das Löschen einzelner Rechnungen,
    nicht das bewusste Entsorgen des ganzen Produkts.
    """
    r = client.post(
        "/api/products",
        json={"name": "Spülmaschine", "kategorie": "Haushalt", "purchase_date": "2026-01-01"},
    )
    assert r.status_code == 201
    product_id = r.json()["id"]

    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    r = client.post(
        "/api/invoices",
        data={"product_id": product_id, "purchase_date": "2026-01-01"},
        files={"file": ("rechnung.png", png_bytes, "image/png")},
    )
    assert r.status_code == 201
    invoice_id = r.json()["id"]
    # Frist läuft noch — Einzellöschung der Rechnung wäre blockiert
    r = client.delete(f"/api/invoices/{invoice_id}")
    assert r.status_code == 409

    # Produkt löschen nimmt die Rechnung trotzdem mit
    r = client.delete(f"/api/products/{product_id}")
    assert r.status_code == 204

    r = client.get(f"/api/invoices/{invoice_id}")
    assert r.status_code == 404
    r = client.get(f"/api/products/{product_id}")
    assert r.status_code == 404



def test_export_xlsx_formula_injection_escaped() -> None:
    """Formelartige Werte dürfen im Excel-Export nie als Formel landen."""
    payload = {
        **_INSURANCE_PAYLOAD,
        "name": '=HYPERLINK("http://evil.example","klick")',
        "vertragsnummer": "=1+1",
        "notes": "@SUM(A1)",
    }
    r = client.post("/api/insurances", json=payload)
    assert r.status_code == 201
    ins_id = r.json()["id"]

    r = client.get("/api/exports/insurances.xlsx")
    assert r.status_code == 200
    wb = load_workbook(BytesIO(r.content))
    ws = wb.active
    values = set()
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            assert cell.data_type != "f", f"Formel-Zelle im Export: {cell.value!r}"
            values.add(cell.value)
    # Inhalt bleibt unverändert als Text erhalten
    assert "=1+1" in values

    client.delete(f"/api/insurances/{ins_id}")


def test_product_link_must_exist() -> None:
    r = client.post(
        "/api/products",
        json={"name": "TV", "kategorie": "Elektronik", "linked_insurance_id": 999999},
    )
    assert r.status_code == 400


def test_upload_corrupt_pdf_returns_400() -> None:
    """Korrupte PDFs (gültige Magic-Bytes, kaputter Inhalt) → 400 statt 500."""
    content = b"%PDF-1.4\n" + b"garbage" * 20
    r = client.post(
        "/api/documents/upload",
        files={"file": ("kaputt.pdf", content, "application/pdf")},
    )
    assert r.status_code == 400


def test_invoice_analyze_corrupt_pdf_returns_400() -> None:
    content = b"%PDF-1.4\n" + b"garbage" * 20
    r = client.post(
        "/api/invoices/analyze",
        files={"file": ("kaputt.pdf", content, "application/pdf")},
    )
    assert r.status_code == 400


def test_attach_list_delete_document() -> None:
    """Dokumente lassen sich an bestehende Versicherungen anhängen, listen und löschen."""
    r = client.post("/api/insurances", json=_INSURANCE_PAYLOAD)
    assert r.status_code == 201
    ins_id = r.json()["id"]

    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    r = client.post(
        f"/api/documents/attach/{ins_id}",
        files={"file": ("beitragsrechnung_2026.png", png_bytes, "image/png")},
    )
    assert r.status_code == 201
    doc = r.json()
    assert doc["insurance_id"] == ins_id
    doc_id = doc["id"]

    r = client.get(f"/api/documents?insurance_id={ins_id}")
    assert r.status_code == 200
    assert any(d["id"] == doc_id for d in r.json())

    r = client.delete(f"/api/documents/{doc_id}")
    assert r.status_code == 204

    r = client.get(f"/api/documents?insurance_id={ins_id}")
    assert all(d["id"] != doc_id for d in r.json())

    client.delete(f"/api/insurances/{ins_id}")


def test_attach_document_insurance_not_found() -> None:
    png_bytes = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    r = client.post(
        "/api/documents/attach/999999",
        files={"file": ("x.png", png_bytes, "image/png")},
    )
    assert r.status_code == 404


def test_delete_insurance_unlinks_products() -> None:
    r = client.post("/api/insurances", json=_INSURANCE_PAYLOAD)
    assert r.status_code == 201
    ins_id = r.json()["id"]

    r = client.post(
        "/api/products",
        json={"name": "Fernseher", "kategorie": "Elektronik", "linked_insurance_id": ins_id},
    )
    assert r.status_code == 201
    product_id = r.json()["id"]

    r = client.delete(f"/api/insurances/{ins_id}")
    assert r.status_code == 204

    r = client.get(f"/api/products/{product_id}")
    assert r.status_code == 200
    assert r.json()["linked_insurance_id"] is None
